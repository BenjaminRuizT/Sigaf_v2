"""Async SQLAlchemy engine + session factory + startup seeding."""
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy import text
import logging, uuid, os, io
from datetime import datetime, timezone
from pathlib import Path

from app.config import settings

logger = logging.getLogger(__name__)

engine = create_async_engine(
    settings.database_url,
    echo=False,
    pool_size=10,
    max_overflow=20,
    pool_pre_ping=True,
)

AsyncSessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

class Base(DeclarativeBase):
    pass

async def get_db():
    async with AsyncSessionLocal() as session:
        try:
            yield session
            await session.commit()
        except Exception:
            await session.rollback()
            raise
        finally:
            await session.close()

async def init_db():
    """Create all tables."""
    from app.models import user_model, equipment_model, audit_model, movement_model  # noqa
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    logger.info("✅ Database tables created/verified")

async def import_initial_data():
    """Seed database from xlsx files if empty."""
    from app.models.equipment_model import Equipment, Store
    from app.models.user_model import User
    from app.security.password_hash import hash_password
    from sqlalchemy import select, func

    async with AsyncSessionLocal() as session:
        count = await session.scalar(select(func.count()).select_from(Equipment))
        if count and count > 0:
            logger.info(f"Data already loaded ({count} equipment records)")
            await _ensure_super_admin(session)
            return

        data_dir = Path(settings.data_dir)
        maf_path = data_dir / "MAF.xlsx"
        users_path = data_dir / "USUARIOS.xlsx"

        if not maf_path.exists():
            logger.warning(f"MAF.xlsx not found at {maf_path}. Skipping import.")
            await _ensure_super_admin(session)
            await session.commit()
            return

        await _import_maf(session, maf_path)
        if users_path.exists():
            await _import_users(session, users_path)
        await _ensure_super_admin(session)
        await session.commit()
        logger.info("✅ Initial data import complete")

async def _import_maf(session, maf_path: Path):
    import openpyxl
    from app.models.equipment_model import Equipment, Store

    logger.info(f"Importing MAF from {maf_path}…")
    wb = openpyxl.load_workbook(maf_path, read_only=True)
    ws = wb.active
    stores_dict = {}
    equipment_list = []
    now = datetime.now(timezone.utc)
    cy, cm = now.year, now.month

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        cr_plaza = str(row[0]) if row[0] else ""
        plaza    = str(row[1]) if row[1] else ""
        cr_tienda = str(row[2]) if row[2] else ""
        tienda   = str(row[3]) if row[3] else ""
        barcode  = str(row[4]).replace('\u202d','').replace('\u202c','').strip() if row[4] else ""
        no_activo = str(int(row[5])) if row[5] else ""
        mes_adq  = int(row[6]) if row[6] else 1
        anio_adq = int(row[7]) if row[7] else 2020
        factura  = str(row[8]).strip() if row[8] else ""
        costo    = float(row[9]) if row[9] else 0.0
        depr     = float(row[10]) if row[10] else 0.0
        vida_util = int(row[11]) if row[11] else 0
        remanente = int(row[12]) if row[12] is not None else 0
        descripcion = str(row[13]) if row[13] else ""
        marca    = str(row[14]) if row[14] else ""
        modelo   = str(row[15]) if row[15] else ""
        serie    = str(row[16]) if row[16] else ""

        meses_trans = (cy - anio_adq) * 12 + (cm - mes_adq)
        vida_restante = max(0, vida_util - meses_trans)
        valor_real = round(max(0.0, costo - depr), 2)

        if cr_tienda not in stores_dict:
            stores_dict[cr_tienda] = Store(
                id=str(uuid.uuid4()), cr_plaza=cr_plaza, plaza=plaza,
                cr_tienda=cr_tienda, tienda=tienda, total_equipment=0, audited=False,
            )
        stores_dict[cr_tienda].total_equipment += 1

        equipment_list.append(Equipment(
            id=str(uuid.uuid4()), cr_plaza=cr_plaza, plaza=plaza,
            cr_tienda=cr_tienda, tienda=tienda, codigo_barras=barcode,
            no_activo=no_activo, mes_adquisicion=mes_adq, anio_adquisicion=anio_adq,
            factura=factura, costo=costo, depreciacion=depr, vida_util=vida_util,
            remanente=remanente, descripcion=descripcion, marca=marca, modelo=modelo,
            serie=serie, meses_transcurridos=meses_trans, vida_util_restante=vida_restante,
            valor_real=valor_real, depreciado=(vida_restante <= 0),
        ))

    wb.close()
    session.add_all(list(stores_dict.values()))
    for i in range(0, len(equipment_list), 2000):
        session.add_all(equipment_list[i:i+2000])
        await session.flush()
    logger.info(f"Imported {len(stores_dict)} stores, {len(equipment_list)} equipment records")

async def _import_users(session, users_path: Path):
    import openpyxl
    from app.models.user_model import User
    from app.security.password_hash import hash_password
    from sqlalchemy import select

    wb = openpyxl.load_workbook(users_path, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        perfil   = str(row[0]).strip()
        nombre   = str(row[1]).strip() if row[1] else ""
        email    = str(row[2]).strip() if row[2] else ""
        password = str(row[3]).strip() if row[3] else "password123"
        existing = await session.scalar(select(User).where(User.email == email))
        if not existing:
            session.add(User(
                id=str(uuid.uuid4()), nombre=nombre, email=email,
                password_hash=hash_password(password), perfil=perfil,
            ))
    wb.close()
    logger.info("Users imported from USUARIOS.xlsx")

async def _ensure_super_admin(session):
    from app.models.user_model import User
    from app.security.password_hash import hash_password
    from sqlalchemy import select

    sa = await session.scalar(select(User).where(User.perfil == "Super Administrador", User.is_backup == False))
    if not sa:
        session.add(User(
            id=str(uuid.uuid4()), nombre="Admin", email="admin@sigaf.com",
            password_hash=hash_password("Admin*1234"), perfil="Super Administrador",
        ))
    # rotating backup admin — credentials = today YYYY-MM-DD
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    backup = await session.scalar(select(User).where(User.is_backup == True))
    if backup:
        backup.email = today
        backup.password_hash = hash_password(today)
    else:
        session.add(User(
            id=str(uuid.uuid4()), nombre="Sistema", email=today,
            password_hash=hash_password(today), perfil="Super Administrador", is_backup=True,
        ))

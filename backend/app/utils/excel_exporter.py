import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE = "1E3C78"
HEADER_BLUE = "2B5BA8"
LIGHT_BLUE = "D6E4F7"
WHITE = "FFFFFF"


def style_title_row(ws, row, text, merge_cols):
    ws.merge_cells(f"A{row}:{get_column_letter(merge_cols)}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    cell.fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def style_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=WHITE),
            right=Side(style="thin", color=WHITE),
            bottom=Side(style="medium", color=WHITE),
        )
    ws.row_dimensions[row].height = 20


def style_data_row(ws, row, values, alt=False):
    fill_color = LIGHT_BLUE if alt else WHITE
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
        cell.font = Font(size=9, name="Calibri")
        cell.border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )
        if isinstance(val, float):
            cell.number_format = '"$"#,##0.00'
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 16


def build_classifications_excel(items: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clasificaciones"
    today = datetime.now().strftime("%d/%m/%Y")
    style_title_row(ws, 1, "SIGAF - Reporte de Clasificaciones de Auditoria", 7)
    ws.cell(row=2, column=1, value=f"Fecha de exportacion: {today}")
    ws.cell(row=2, column=1).font = Font(size=9, italic=True, name="Calibri", color="666666")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 15
    style_header_row(ws, 3, ["Fecha", "Codigo Barras", "Clasificacion", "Descripcion", "Marca", "Modelo", "Tienda"])
    for i, item in enumerate(items):
        eq = item.get("equipment_data") or {}
        style_data_row(ws, i + 4, [
            (item.get("scanned_at", "") or "")[:19].replace("T", " "),
            item.get("codigo_barras", ""),
            item.get("classification", ""),
            eq.get("descripcion", ""),
            eq.get("marca", ""),
            eq.get("modelo", ""),
            eq.get("tienda", ""),
        ], alt=(i % 2 == 0))
    for col, width in zip(range(1, 8), [18, 14, 16, 30, 14, 16, 24]):
        ws.column_dimensions[get_column_letter(col)].width = width
    _add_footer(ws)
    return _save(wb)


def build_movements_excel(items: list[dict], export_type: str, plaza: str = "") -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    today_display = datetime.now().strftime("%d/%m/%Y")

    if export_type == "movements-ab":
        ws.title = "Altas y Bajas"
        doc_title = "Formato de Movimiento de AF — ALTAS y BAJAS"
        headers = ["Plaza Origen", "Tipo de Movimiento", "Numero de Activo", "Codigo de Barras",
                   "Descripcion del Equipo", "Valor Real", "Marca", "Modelo", "Anyo",
                   "Numero de Serie", "CR Tienda Origen", "Tienda Origen"]
        merge_cols = 12
    elif export_type == "movements-transferencias":
        ws.title = "Transferencias"
        doc_title = "Formato de Movimiento de AF — TRANSFERENCIAS"
        headers = ["Plaza Destino", "Tipo de Movimiento", "Numero de Activo", "Codigo de Barras",
                   "Descripcion del Equipo", "Valor Real", "Marca", "Modelo", "Anyo",
                   "Numero de Serie", "CR Tienda Origen", "Tienda Origen", "CR Tienda Destino", "Tienda Destino"]
        merge_cols = 14
    else:
        ws.title = "Movimientos"
        doc_title = "Formato de Movimiento de AF"
        headers = ["Plaza", "Tipo de Movimiento", "Numero de Activo", "Codigo de Barras",
                   "Descripcion del Equipo", "Valor Real", "Marca", "Modelo", "Anyo",
                   "Numero de Serie", "CR Tienda Origen", "Tienda Origen", "CR Tienda Destino", "Tienda Destino"]
        merge_cols = 14

    style_title_row(ws, 1, doc_title, merge_cols)
    ws.cell(row=2, column=1, value="FECHA:").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=2, column=2, value=today_display).font = Font(size=10, name="Calibri")
    ws.cell(row=2, column=4, value="DEPARTAMENTO:").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=2, column=5, value="Sistemas").font = Font(size=10, name="Calibri")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 5
    style_header_row(ws, 4, headers)

    tipo_labels = {"alta": "ALTA", "baja": "BAJA", "disposal": "BAJA", "transfer": "TRANSFERENCIA"}
    for i, item in enumerate(items):
        eq = item.get("equipment_data") or {}
        tipo = tipo_labels.get(item.get("type", ""), (item.get("type", "") or "").upper())
        row_vals = [
            item.get("plaza", "") or eq.get("plaza", ""),
            tipo,
            eq.get("no_activo", ""),
            eq.get("codigo_barras", ""),
            eq.get("descripcion", ""),
            float(eq.get("valor_real", 0) or 0),
            eq.get("marca", ""),
            eq.get("modelo", ""),
            eq.get("año_adquisicion", ""),
            eq.get("serie", ""),
            item.get("from_cr_tienda", ""),
            item.get("from_tienda", ""),
        ]
        if merge_cols >= 14:
            row_vals += [item.get("to_cr_tienda", ""), item.get("to_tienda", "")]
        style_data_row(ws, i + 5, row_vals, alt=(i % 2 == 0))

    col_widths = [16, 16, 14, 14, 34, 12, 14, 16, 8, 18, 14, 26, 14, 26]
    for col, width in enumerate(col_widths[:merge_cols], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _add_footer(ws)
    return _save(wb)


def build_audits_excel(items: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditorias"
    today = datetime.now().strftime("%d/%m/%Y")
    style_title_row(ws, 1, "SIGAF - Historial de Auditorias", 13)
    ws.cell(row=2, column=1, value=f"Fecha de exportacion: {today}")
    ws.cell(row=2, column=1).font = Font(size=9, italic=True, name="Calibri", color="666666")
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 15
    style_header_row(ws, 3, ["Fecha Inicio", "Fecha Fin", "Tienda", "CR", "Plaza", "Auditor",
                               "Estado", "Total Equipos", "Localizados", "Sobrantes",
                               "No Localizados", "Valor No Localizado", "Motivo Cancelacion"])
    for i, item in enumerate(items):
        style_data_row(ws, i + 4, [
            (item.get("started_at", "") or "")[:19].replace("T", " "),
            (item.get("finished_at", "") or "")[:19].replace("T", " "),
            item.get("tienda", ""), item.get("cr_tienda", ""), item.get("plaza", ""),
            item.get("auditor_name", ""), item.get("status", ""),
            item.get("total_equipment", 0), item.get("located_count", 0), item.get("surplus_count", 0),
            item.get("not_found_count", 0), float(item.get("not_found_value", 0) or 0),
            item.get("cancel_reason", ""),
        ], alt=(i % 2 == 0))
    for col, width in zip(range(1, 14), [18, 18, 28, 10, 14, 20, 12, 12, 12, 12, 14, 16, 28]):
        ws.column_dimensions[get_column_letter(col)].width = width
    _add_footer(ws)
    return _save(wb)


def _add_footer(ws):
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value=f"Generado por SIGAF — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=last_row, column=1).font = Font(size=8, italic=True, color="999999", name="Calibri")


def _save(wb: openpyxl.Workbook) -> io.BytesIO:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

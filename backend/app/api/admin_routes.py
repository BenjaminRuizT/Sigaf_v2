"""Admin routes — user management, equipment edit, data reset, export, templates, PDF downloads."""
import io, uuid, json
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from app.database import get_db, _import_maf, _import_users, _ensure_super_admin
from app.models.user_model import User
from app.models.equipment_model import Equipment, Store
from app.models.audit_model import Audit, AuditScan
from app.models.movement_model import Movement
from app.security.jwt_handler import get_current_user, require_super_admin
from app.security.password_hash import hash_password

router = APIRouter(tags=["admin"])


class UserCreateInput(BaseModel):
    nombre: str; email: str; password: str; perfil: str

class UserUpdateInput(BaseModel):
    nombre: Optional[str]=None; email: Optional[str]=None; password: Optional[str]=None; perfil: Optional[str]=None

class EquipmentUpdateInput(BaseModel):
    codigo_barras: Optional[str]=None; no_activo: Optional[str]=None
    descripcion: Optional[str]=None; marca: Optional[str]=None
    modelo: Optional[str]=None; serie: Optional[str]=None
    costo: Optional[float]=None; depreciacion: Optional[float]=None


# ── Users ─────────────────────────────────────────────────────────────────────
@router.get("/admin/users")
async def get_users(user=Depends(require_super_admin), db: AsyncSession = Depends(get_db)):
    users = (await db.scalars(select(User).where(User.is_backup == False))).all()
    return [u.to_dict() for u in users]

@router.post("/admin/users")
async def create_user(body: UserCreateInput, user=Depends(require_super_admin), db: AsyncSession = Depends(get_db)):
    existing = await db.scalar(select(User).where(User.email == body.email))
    if existing: raise HTTPException(400, "Email already exists")
    u = User(id=str(uuid.uuid4()), nombre=body.nombre, email=body.email,
             password_hash=hash_password(body.password), perfil=body.perfil)
    db.add(u); await db.flush()
    return u.to_dict()

@router.put("/admin/users/{user_id}")
async def update_user(user_id: str, body: UserUpdateInput, user=Depends(require_super_admin), db: AsyncSession = Depends(get_db)):
    u = await db.get(User, user_id)
    if not u: raise HTTPException(404, "User not found")
    if body.nombre:   u.nombre = body.nombre
    if body.email:    u.email = body.email
    if body.password: u.password_hash = hash_password(body.password)
    if body.perfil:   u.perfil = body.perfil
    await db.flush()
    return u.to_dict()

@router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, user=Depends(require_super_admin), db: AsyncSession = Depends(get_db)):
    if user_id == user["sub"]: raise HTTPException(400, "No puede eliminarse a si mismo")
    u = await db.get(User, user_id)
    if not u: raise HTTPException(404, "User not found")
    if u.is_backup: raise HTTPException(400, "No se puede eliminar este usuario")
    await db.delete(u); await db.flush()
    return {"message": "User deleted"}

# ── Equipment ─────────────────────────────────────────────────────────────────
@router.get("/admin/equipment")
async def admin_get_equipment(
    search: Optional[str]=None, plaza: Optional[str]=None, page: int=1, limit: int=50,
    user=Depends(require_super_admin), db: AsyncSession = Depends(get_db),
):
    q = select(Equipment)
    if plaza and plaza != "all": q = q.where(Equipment.plaza == plaza)
    if search:
        q = q.where(or_(Equipment.codigo_barras.ilike(f"%{search}%"), Equipment.no_activo.ilike(f"%{search}%"),
                        Equipment.descripcion.ilike(f"%{search}%"), Equipment.tienda.ilike(f"%{search}%")))
    total = await db.scalar(select(func.count()).select_from(q.subquery()))
    items = (await db.scalars(q.offset((page-1)*limit).limit(limit))).all()
    return {"items": [e.to_dict() for e in items], "total": total, "page": page, "pages": max(1, -(-total // limit))}

@router.put("/admin/equipment/{equip_id}")
async def admin_update_equipment(equip_id: str, body: EquipmentUpdateInput, user=Depends(require_super_admin), db: AsyncSession = Depends(get_db)):
    eq = await db.get(Equipment, equip_id)
    if not eq: raise HTTPException(404, "Equipment not found")
    for field in ("codigo_barras","no_activo","descripcion","marca","modelo","serie"):
        v = getattr(body, field)
        if v is not None: setattr(eq, field, v)
    if body.costo is not None:       eq.costo = body.costo
    if body.depreciacion is not None: eq.depreciacion = body.depreciacion
    eq.valor_real = round(max(0, eq.costo - eq.depreciacion), 2)
    await db.flush()
    return eq.to_dict()

# ── Stores ────────────────────────────────────────────────────────────────────
@router.get("/admin/stores")
async def admin_get_stores(
    search: Optional[str]=None, plaza: Optional[str]=None, page: int=1, limit: int=50,
    user=Depends(require_super_admin), db: AsyncSession = Depends(get_db),
):
    q = select(Store)
    if plaza and plaza != "all": q = q.where(Store.plaza == plaza)
    if search:
        q = q.where(or_(Store.cr_tienda.ilike(f"%{search}%"), Store.tienda.ilike(f"%{search}%")))
    total = await db.scalar(select(func.count()).select_from(q.subquery()))
    items = (await db.scalars(q.order_by(Store.tienda).offset((page-1)*limit).limit(limit))).all()
    return {"items": [s.to_dict() for s in items], "total": total, "page": page, "pages": max(1, -(-total // limit))}

# ── Templates ─────────────────────────────────────────────────────────────────
@router.get("/admin/template/{template_type}")
async def download_template(template_type: str, user=Depends(require_super_admin)):
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    if template_type == "maf":
        ws.title = "MAF"
        ws.append(["Cr Plaza","Plaza","Cr Tienda","Tienda","Codigo Barras","No Activo",
                   "Mes Adquisicion","Año Adquisicion","Factura","Costo","Depresiacion",
                   "Vida util","Remanente","Descripción","Marca","Modelo","Serie"])
        ws.append(["32ECK","Este","31DYQ","Administración TIJ Este","04001201","6950216",
                   "9","2021","108771","9400","9306","40","0","IMPRESORA","EPSON","FX890II","X3YF049071"])
    elif template_type == "usuarios":
        ws.title = "USUARIOS"
        ws.append(["Perfil","Nombre","Email","Contraseña"])
        ws.append(["Super Administrador","Juan Pérez","juan.perez@empresa.com","MiContraseña*1"])
        ws.append(["Administrador","María López","maria.lopez@empresa.com","MiContraseña*1"])
    else:
        raise HTTPException(400, "Invalid template type")
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=template_{template_type}.xlsx"})

# ── Data Reset ────────────────────────────────────────────────────────────────
MAF_EXPECTED_HEADERS = ["cr plaza","plaza","cr tienda","tienda","codigo barras","no activo",
    "mes adquisicion","año adquisicion","factura","costo","depresiacion","vida util","remanente","descripción","marca","modelo","serie"]

@router.post("/admin/reset-data")
async def reset_data(maf_file: UploadFile = File(...), users_file: UploadFile = File(...),
                     user=Depends(require_super_admin), db: AsyncSession = Depends(get_db)):
    import openpyxl
    from pathlib import Path; from app.config import settings

    if not maf_file.filename.endswith('.xlsx') or not users_file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Ambos archivos deben ser formato .xlsx")

    maf_content   = await maf_file.read()
    users_content = await users_file.read()

    # Validate MAF
    try:
        wb = openpyxl.load_workbook(io.BytesIO(maf_content), read_only=True)
        ws = wb.active
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headers: raise ValueError("MAF.xlsx: Archivo vacío")
        actual = [str(h).strip().lower() if h else "" for h in headers]
        if len(actual) < 17: raise ValueError(f"MAF.xlsx: Se requieren 17 columnas, encontradas {len(actual)}")
        wb.close()
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Save files
    data_dir = Path(settings.data_dir); data_dir.mkdir(parents=True, exist_ok=True)
    (data_dir / "MAF.xlsx").write_bytes(maf_content)
    (data_dir / "USUARIOS.xlsx").write_bytes(users_content)

    # Drop and reimport
    for cls in (AuditScan, Movement, Audit, Equipment, Store, User):
        for obj in (await db.scalars(select(cls))).all():
            await db.delete(obj)
    await db.flush()

    from io import BytesIO
    from pathlib import Path as P
    await _import_maf(db, P(settings.data_dir) / "MAF.xlsx")
    await _import_users(db, P(settings.data_dir) / "USUARIOS.xlsx")
    await _ensure_super_admin(db)
    await db.flush()

    count_eq     = await db.scalar(select(func.count()).select_from(Equipment))
    count_stores = await db.scalar(select(func.count()).select_from(Store))
    count_users  = await db.scalar(select(func.count()).select_from(User).where(User.is_backup == False))
    return {"message": "Data reset complete", "equipment": count_eq, "stores": count_stores, "users": count_users}

# ── Export to Excel ────────────────────────────────────────────────────────────
@router.get("/export/{export_type}")
async def export_excel(
    export_type: str,
    token: Optional[str] = None, classification: Optional[str] = None,
    type: Optional[str] = None, status: Optional[str] = None,
    search: Optional[str] = None, plaza: Optional[str] = None,
    user=Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK_BLUE="1E3C78"; HEADER_BLUE="2B5BA8"; LIGHT_BLUE="D6E4F7"; WHITE="FFFFFF"
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_display = datetime.now().strftime("%d/%m/%Y")

    def hrow(ws, row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            c.fill = PatternFill(fill_type="solid", fgColor=HEADER_BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20

    def drow(ws, row, values, alt=False):
        fill_color = LIGHT_BLUE if alt else WHITE
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 16

    wb = openpyxl.Workbook(); ws = wb.active

    if export_type == "classifications":
        ws.title = "Clasificaciones"
        q = select(AuditScan)
        if classification and classification != "all": q = q.where(AuditScan.classification == classification)
        items = (await db.scalars(q.order_by(AuditScan.scanned_at.desc()))).all()
        ws.cell(1,1,"SIGAF - Reporte de Clasificaciones").font = Font(bold=True, size=14)
        hrow(ws, 2, ["Fecha","Código Barras","Clasificación","Descripción","Marca","Modelo","Tienda"])
        for i, s in enumerate(items):
            eq = s.to_dict().get("equipment_data") or {}
            drow(ws, i+3, [
                s.scanned_at.strftime("%Y-%m-%d %H:%M") if s.scanned_at else "",
                s.codigo_barras, s.classification,
                eq.get("descripcion",""), eq.get("marca",""), eq.get("modelo",""), eq.get("tienda",""),
            ], alt=i%2==0)
        filename = f"sigaf_classifications_{today_str}.xlsx"

    elif export_type == "audits":
        ws.title = "Auditorías"
        q = select(Audit)
        if status and status != "all": q = q.where(Audit.status == status)
        if search: q = q.where(or_(Audit.tienda.ilike(f"%{search}%"), Audit.cr_tienda.ilike(f"%{search}%")))
        items = (await db.scalars(q.order_by(Audit.started_at.desc()))).all()
        hrow(ws, 1, ["Inicio","Fin","Tienda","CR","Plaza","Auditor","Estado","Total","Localizados","Sobrantes","No Loc.","Valor NL","Cancelación"])
        for i, a in enumerate(items):
            drow(ws, i+2, [
                a.started_at.strftime("%Y-%m-%d %H:%M") if a.started_at else "",
                a.finished_at.strftime("%Y-%m-%d %H:%M") if a.finished_at else "",
                a.tienda, a.cr_tienda, a.plaza, a.auditor_name, a.status,
                a.total_equipment, a.located_count, a.surplus_count, a.not_found_count,
                a.not_found_value, a.cancel_reason or "",
            ], alt=i%2==0)
        filename = f"sigaf_audits_{today_str}.xlsx"

    elif export_type in ("movements","movements-ab","movements-transferencias"):
        ws.title = "Movimientos"
        q = select(Movement)
        if export_type == "movements-ab":
            q = q.where(Movement.type.in_(["alta","baja","disposal"]))
        elif export_type == "movements-transferencias":
            q = q.where(Movement.type == "transfer")
        if plaza and plaza != "all": q = q.where(Movement.plaza == plaza)
        items = (await db.scalars(q.order_by(Movement.created_at.desc()))).all()
        hrow(ws, 1, ["Plaza","Tipo","No Activo","Código Barras","Descripción","Valor Real","Marca","Modelo","Año","Serie","CR Origen","Tienda Origen","CR Destino","Tienda Destino"])
        tipo_labels = {"alta":"ALTA","baja":"BAJA","disposal":"BAJA","transfer":"TRANSFERENCIA"}
        for i, m in enumerate(items):
            eq = m.to_dict().get("equipment_data") or {}
            drow(ws, i+2, [
                m.plaza or "", tipo_labels.get(m.type, m.type.upper()),
                eq.get("no_activo",""), eq.get("codigo_barras",""), eq.get("descripcion",""),
                eq.get("valor_real",0), eq.get("marca",""), eq.get("modelo",""),
                eq.get("anio_adquisicion",""), eq.get("serie",""),
                m.from_cr_tienda or "", m.from_tienda or "",
                m.to_cr_tienda or "", m.to_tienda or "",
            ], alt=i%2==0)
        filename = f"sigaf_{export_type}_{today_str}.xlsx"
    else:
        raise HTTPException(400, "Invalid export type")

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

# ── PDF Downloads ──────────────────────────────────────────────────────────────
@router.get("/download/manual")
async def download_manual(user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from app.utils.pdf_generator import generate_user_manual
    stats = await _get_stats(db)
    out = generate_user_manual(stats, [])
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=SIGAF_Manual.pdf"})

@router.get("/download/presentation")
async def download_presentation(user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from app.utils.pdf_generator import generate_presentation
    stats = await _get_stats(db)
    out = generate_presentation(stats, [])
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=SIGAF_Presentacion.pdf"})

async def _get_stats(db):
    total_stores = await db.scalar(select(func.count()).select_from(Store))
    audited      = await db.scalar(select(func.count()).select_from(Store).where(Store.audited == True))
    total_eq     = await db.scalar(select(func.count()).select_from(Equipment))
    depr_eq      = await db.scalar(select(func.count()).select_from(Equipment).where(Equipment.depreciado == True))
    total_cost   = await db.scalar(select(func.sum(Equipment.costo))) or 0
    total_real   = await db.scalar(select(func.sum(Equipment.valor_real))) or 0
    completed    = await db.scalar(select(func.count()).select_from(Audit).where(Audit.status == "completed"))
    plaza_rows   = (await db.execute(select(Equipment.plaza, func.count(Equipment.id)).group_by(Equipment.plaza))).all()
    return {
        "total_stores": total_stores, "audited_stores": audited,
        "total_equipment": total_eq, "deprecated_equipment": depr_eq,
        "active_equipment": total_eq - depr_eq, "total_cost": total_cost,
        "total_real_value": total_real, "completed_audits": completed,
        "equipment_by_plaza": {r[0]: r[1] for r in plaza_rows if r[0]},
    }
